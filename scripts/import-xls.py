#!/usr/bin/env python3
"""
import-xls.py — 航旅纵横 XLS/XLSX → flights.json 导入工具

用法:
    # 查看文件结构
    python3 scripts/import-xls.py data/exports/文件.xls --inspect

    # 导入指定工作表 (默认 sheet 0)
    python3 scripts/import-xls.py data/exports/文件.xls --sheet 1
    python3 scripts/import-xls.py data/exports/文件.xls --sheet "已结束"
"""

import json
import os
import re
import sys
from pathlib import Path


def read_xlsx_rows(filepath, sheet_idx=0):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_idx]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def read_xls_rows(filepath, sheet_idx=0):
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(sheet_idx)
    rows = []
    for r in range(ws.nrows):
        rows.append(tuple(ws.cell_value(r, c) for c in range(ws.ncols)))
    return rows


COLUMN_MAP = {
    '日期': 'date', 'date': 'date', '飞行日期': 'date',
    '航空公司': 'airline', '航司': 'airline', 'airline': 'airline',
    '航班号': 'flightNo', '航班': 'flightNo', 'flight_no': 'flightNo',
    '出发城市': 'departureCity', '出发地': 'departureCity', '出发': 'departureCity',
    '出发时间': 'depTime', '起飞时间': 'depTime', '起飞': 'depTime',
    '到达城市': 'arrivalCity', '到达地': 'arrivalCity', '到达': 'arrivalCity',
    '到达时间': 'arrTime', 'arr_time': 'arrTime',
    '里程数': 'distance', '里程': 'distance', 'distance': 'distance',
    '客票号': 'ticketNo', '票号': 'ticketNo',
    '客票状态': 'status', '状态': 'status',
    '机型': 'aircraft', 'aircraft': 'aircraft',
    '座位号': 'seat', '座位': 'seat', 'seat': 'seat',
    '备注': 'notes', 'note': 'notes',
}


def parse_date_cn(s):
    """2026年7月2日 → 2026-07-02"""
    s = str(s).strip()
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日?', s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m2 = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m2:
        return s[:10]
    return s


def parse_distance(d):
    s = str(d).strip()
    m = re.search(r'(\d+)', s)
    return int(m.group(1)) if m else 0


def auto_map_columns(header):
    mapping = {}
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        key = str(cell).strip()
        if key in COLUMN_MAP:
            mapping[idx] = COLUMN_MAP[key]
        else:
            for alias, field in COLUMN_MAP.items():
                if alias in key or key[:2] == alias[:2]:
                    mapping[idx] = field
                    break
    return mapping


def resolve_sheet(filepath, spec):
    """解析 --sheet 参数为 sheet index"""
    if spec is None:
        return 0
    try:
        return int(spec)
    except ValueError:
        pass
    import xlrd
    wb = xlrd.open_workbook(filepath)
    for i in range(wb.nsheets):
        if spec.lower() in wb.sheet_names()[i].lower():
            return i
    print(f"未找到匹配的工作表: {spec}")
    sys.exit(1)


def inspect(filepath, sheet_idx=0):
    rows = read_xls_rows(filepath, sheet_idx) if filepath.endswith('.xls') else read_xlsx_rows(filepath, sheet_idx)
    import xlrd
    wb = xlrd.open_workbook(filepath)
    sheet_name = wb.sheet_names()[sheet_idx]

    print(f"📋 工作表 [{sheet_idx}]: {sheet_name}")
    print(f"   总行数: {len(rows)}, 列数: {len(rows[0]) if rows else 0}\n")
    for i, row in enumerate(rows[:6]):
        vals = [str(v)[:35] if v is not None else '' for v in row]
        print(f"  行{i+1}: {' │ '.join(vals)}")
    print()

    if len(rows) > 0:
        mapping = auto_map_columns(rows[0])
        unmapped = [str(rows[0][i])[:25] for i in range(len(rows[0])) if rows[0][i] and i not in mapping]
        if unmapped:
            print(f"⚠ 未映射列: {', '.join(unmapped)}")
        print(f"✓ 已映射 {len(mapping)}/{len([c for c in rows[0] if c])} 列")


def run_import(filepath, output_path, sheet_idx=0, status_filter=None):
    rows = read_xls_rows(filepath, sheet_idx) if filepath.endswith('.xls') else read_xlsx_rows(filepath, sheet_idx)
    if len(rows) < 2:
        print("文件没有数据")
        return

    header = rows[0]
    mapping = auto_map_columns(header)
    unmapped = [str(header[i])[:25] for i in range(len(header)) if header[i] and i not in mapping]
    if unmapped:
        print(f"⚠ 未映射列: {', '.join(unmapped)}")

    new_flights = []
    for row_idx, row in enumerate(rows[1:], 2):
        row_data = {}
        for col_idx, field in mapping.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip():
                row_data[field] = str(val).strip()
        if 'flightNo' not in row_data or 'date' not in row_data:
            continue

        # Filter by status if specified
        if status_filter and row_data.get('status', '') not in status_filter:
            continue

        row_data['date'] = parse_date_cn(row_data.get('date', ''))
        if 'distance' in row_data:
            row_data['distance'] = parse_distance(row_data['distance'])

        fid = row_data['flightNo']
        date_short = row_data['date'].replace('-', '')[-4:] if '-' in row_data.get('date', '') else ''
        row_data['id'] = f"{fid}-{date_short}" if date_short else fid
        row_data.setdefault('airlineEn', row_data.get('airline', ''))
        row_data.setdefault('airlineCode', fid[:2] if fid else '')
        row_data.setdefault('cabin', 'Y')
        row_data.setdefault('seat', '')
        row_data.setdefault('aircraft', '')
        row_data.setdefault('tripId', '')
        row_data.setdefault('notes', '')
        row_data.setdefault('duration', '')
        row_data.setdefault('departure', '')
        row_data.setdefault('arrival', '')
        row_data.setdefault('departureCityEn', row_data.get('departureCity', ''))
        row_data.setdefault('arrivalCityEn', row_data.get('arrivalCity', ''))

        new_flights.append(row_data)

    print(f"解析出 {len(new_flights)} 条航班")

    if output_path:
        if os.path.exists(output_path):
            with open(output_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
        else:
            existing = {'flights': []}

        existing_ids = {f['id'] for f in existing['flights']}
        added = 0
        for f in new_flights:
            if f['id'] not in existing_ids:
                existing['flights'].append(f)
                existing_ids.add(f['id'])
                added += 1

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)

        print(f"✅ 已合并 → {output_path}")
        print(f"   新增 {added} 条，现有共 {len(existing['flights'])} 条航班")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        sys.exit(1)

    script_dir = Path(__file__).parent.parent
    output_path = script_dir / 'data' / 'flights.json'

    # Parse args
    inspect_only = '--inspect' in sys.argv
    sheet_spec = None
    if '--sheet' in sys.argv:
        idx = sys.argv.index('--sheet')
        if idx + 1 < len(sys.argv):
            sheet_spec = sys.argv[idx + 1]

    sheet_idx = resolve_sheet(filepath, sheet_spec)

    if inspect_only:
        inspect(filepath, sheet_idx)
    else:
        run_import(filepath, output_path, sheet_idx)
